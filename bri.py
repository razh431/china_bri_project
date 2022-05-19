import os
import openpyxl
import xlrd
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

import functools

from collections import OrderedDict

plt.style.use('dark_background')

# import required module


def addlabels(x, y):
    for i in range(len(x)):
        plt.text(i, y[i], y[i])


def find_specific_cell(data_val, c_or_v, currentSheet):
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEFGHIJKL":
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == data_val:
                # print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                total = "{}{}".format(c_or_v, row)
                # print("Owes {} a total of {}".format(
                #     currentSheet[cell_name].value, currentSheet[total].value))
                return currentSheet[total].value


def plot_top(arr):
    print(arr)
    arr = arr[::-1]

    a = np.array([year['debt'] for year in arr])
    for i, value in zip(np.linspace(-0.2, 0.2, 5), a.T):
        print(np.arange(9)+i, value)
        plt.bar(np.arange(9)+i, value, width=0.1)
    plt.xticks(np.arange(9), [2013 + i for i in range(9)])
    plt.legend(['NETHERLANDS', 'UNITED KINGDOM',
                'UNITED STATES OF AMERICA', 'CHINA', 'FRANCE'])
    plt.title("Debt from 2013-2021")
    plt.show()


def top_countries():
    years = []
    debt_owed = []
    # iterate over files in
    # that directory

    countries = []
    countries_debt = []

    path = os.listdir(
        "/Users/rachelzhou/Documents/Sp 2022/china/kazakhstan_data")
    path.sort()
    # print(path[0])

    # for filename in range(1):
    #     filename = "1 quarter 2020.xlsx"
    for filename in path:
        top_countries = {'country': [], 'debt': []}
        f = os.path.join(
            "/Users/rachelzhou/Documents/Sp 2022/china/kazakhstan_data/", filename)
        print(filename)
        years.append(int(f[-9:-5]))

        theFile = openpyxl.load_workbook(f)
        allSheetNames = theFile.sheetnames

        df = pd.read_excel(f, sheet_name="1. country")
        top = df.iloc[8:].sort_values(df.columns[1], ascending=False).head(5)

        for i in range(len(top)):
            debt_val = top.iloc[i, 1]
            currentSheet = theFile["1. country"]

            country = (find_specific_cell(debt_val, "A", currentSheet))

            top_countries['country'].append(country)

            top_countries['debt'].append(debt_val)

            # print(str(country) + " " + str(debt_val))
        print(top_countries)
        res = {top_countries['country'][i]: top_countries['debt'][i]
               for i in range(len(top_countries['country']))}
        # print(res)
        keyorder = ["NETHERLANDS", "UNITED KINGDOM", "UNITED STATES OF AMERICA",
                    "CHINA", "FRANCE", "Not determined by country***", "Not determined by country*", "INTERNATIONAL ORGANIZATIONS"]

        a = {k: res[k] for k in keyorder if k in res}

        countries = []
        debt = []
        for key, value in a.items():
            countries.append(key)
            debt.append(value)
        top_countries = {"country": countries, "debt": debt}
        print(top_countries)
        # print({k: top_countries[k] for k in keyorder if k in top_countries})
        countries_debt.append(top_countries)

    print(countries_debt)

    plot_top(countries_debt)


def specific_country():

    years = []
    debt_owed = []

    path = os.listdir(
        "/Users/rachelzhou/Documents/Sp 2022/china/kazakhstan_data")
    path.sort()

    for filename in path:
        f = os.path.join(
            "/Users/rachelzhou/Documents/Sp 2022/china/kazakhstan_data/", filename)
        years.append(int(f[-9:-5]))
        theFile = openpyxl.load_workbook(f)
        currentSheet = theFile["1. country"]
        debt = (find_specific_cell("NETHERLANDS", "B", currentSheet))
        debt_owed.append(debt)

    Data = {'Year': years,
            'debt': debt_owed
            }

    print(years)
    print(debt_owed)

    df = pd.DataFrame(Data, columns=['Year', 'debt'])

    plt.plot(df['Year'], df['debt'], color='red', marker='o')

    debt_trun = [int(x) for x in df['debt']]

    addlabels(df['Year'], debt_trun)

    plt.title('Debt Owed to Netherlands')
    plt.xlabel('year')
    plt.ylabel('Debt Owed (millions USD)')

    plt.grid(True)
    plt.show()


# specific_country()
top_countries()
